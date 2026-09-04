import io
import uuid
import base64
import urllib.parse
from datetime import datetime
from typing import Dict, Any, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Generate a unique human-friendly ticket code
def generate_ticket_code(event_id: int) -> str:
    short_uuid = uuid.uuid4().hex[:6].upper()
    return f"REG-E{event_id}-{short_uuid}"

# Generate QR code as Base64 Data URI
def generate_qr_base64(data: str) -> str:
    try:
        import qrcode
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10,
            border=2,
        )
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="#1e293b", back_color="#ffffff")
        
        buffer = io.BytesIO()
        img.save(buffer, format="PNG")
        b64_encoded = base64.b64encode(buffer.getvalue()).decode("utf-8")
        return f"data:image/png;base64,{b64_encoded}"
    except Exception as e:
        # Fallback if qrcode library is not yet ready
        return ""

# Generate WhatsApp direct link
def generate_whatsapp_link(phone: str, attendee_name: str, event_title: str, template_type: str = "missed") -> str:
    if not phone:
        return ""
    
    # Strip spaces, dashes, parentheses
    clean_phone = "".join(c for c in phone if c.isdigit() or c == "+")
    if clean_phone.startswith("+"):
        clean_phone = clean_phone[1:]
    
    if template_type == "missed":
        message = (
            f"Hello {attendee_name},\n\n"
            f"We noticed that you were unable to attend '{event_title}' today. "
            f"We truly missed having you with us! Could you please let us know if everything is alright or if you faced any issues joining?\n\n"
            f"Best regards,\nEvent Organizing Team"
        )
    elif template_type == "materials":
        message = (
            f"Hello {attendee_name},\n\n"
            f"Thank you for registering for '{event_title}'. Although you couldn't make it in person, "
            f"we're happy to share the presentation slides and event summary with you.\n\n"
            f"Feel free to reply if you have any questions!\n\n"
            f"Warm regards,\nEvent Organizing Team"
        )
    else:
        message = (
            f"Hello {attendee_name},\n\n"
            f"We hope you are doing well! Following up regarding your registration for '{event_title}'. "
            f"Please let us know how we can assist you for upcoming sessions.\n\n"
            f"Regards,\nEvent Organizing Team"
        )
        
    encoded_message = urllib.parse.quote(message)
    return f"https://wa.me/{clean_phone}?text={encoded_message}"

# Export event report to Excel (.xlsx) with styled sheets
def generate_excel_report(event: Dict[str, Any], attendees: List[Dict[str, Any]], followups: List[Dict[str, Any]]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    
    # Fonts and Fills
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
    sub_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid") # Blue
    kpi_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="1E3A8A")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # -------------------------------------------------------------
    # SHEET 1: Executive Summary
    # -------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    ws_summary["A1"] = f"Event Final Report: {event.get('title', 'Event')}"
    ws_summary["A1"].font = title_font
    
    # Event Info Box
    info_rows = [
        ("Event Date:", event.get("event_date", "")),
        ("Time:", f"{event.get('start_time', '')} - {event.get('end_time', '')}"),
        ("Venue / Location:", f"{event.get('venue', '')} ({event.get('location_type', '')})"),
        ("Category:", event.get("category", "")),
        ("Max Capacity:", event.get("max_capacity", 0)),
        ("Report Generated At:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    
    for r_idx, (label, val) in enumerate(info_rows, start=3):
        ws_summary[f"A{r_idx}"] = label
        ws_summary[f"A{r_idx}"].font = bold_font
        ws_summary[f"B{r_idx}"] = val
        ws_summary[f"B{r_idx}"].font = regular_font
    
    # Calculate KPIs
    total_registered = len(attendees)
    attended_count = sum(1 for a in attendees if a.get("status") == "Attended")
    absent_count = sum(1 for a in attendees if a.get("status") == "Registered")
    walkin_count = sum(1 for a in attendees if a.get("is_walkin") == 1)
    attendance_rate = round((attended_count / total_registered * 100), 1) if total_registered > 0 else 0
    
    # KPI Table Header
    ws_summary["A10"] = "Key Performance Metrics"
    ws_summary["A10"].font = Font(name="Calibri", size=13, bold=True, color="1E3A8A")
    
    kpi_headers = ["Metric", "Count", "Percentage"]
    for col_idx, h in enumerate(kpi_headers, start=1):
        cell = ws_summary.cell(row=11, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    metrics = [
        ("Total Registrations", total_registered, "100%"),
        ("Attended (Checked In)", attended_count, f"{attendance_rate}%"),
        ("No-Shows / Absent", absent_count, f"{round((absent_count / total_registered * 100), 1) if total_registered > 0 else 0}%"),
        ("Walk-in Attendances", walkin_count, f"{round((walkin_count / total_registered * 100), 1) if total_registered > 0 else 0}%"),
    ]
    
    for r_idx, (m_label, m_val, m_pct) in enumerate(metrics, start=12):
        c1 = ws_summary.cell(row=r_idx, column=1, value=m_label)
        c2 = ws_summary.cell(row=r_idx, column=2, value=m_val)
        c3 = ws_summary.cell(row=r_idx, column=3, value=m_pct)
        for c in (c1, c2, c3):
            c.border = thin_border
            c.font = regular_font
        c2.alignment = Alignment(horizontal="center")
        c3.alignment = Alignment(horizontal="center")
        
    # Auto-fit columns
    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 35
    ws_summary.column_dimensions["C"].width = 18

    # -------------------------------------------------------------
    # Parse Custom Field Configuration if any
    # -------------------------------------------------------------
    import json
    custom_fields = []
    if event.get("custom_fields_config"):
        try:
            cfg = json.loads(event["custom_fields_config"]) if isinstance(event["custom_fields_config"], str) else event["custom_fields_config"]
            custom_fields = [f for f in cfg if not f.get("is_standard") and f.get("enabled", True)]
        except Exception:
            custom_fields = []

    # -------------------------------------------------------------
    # SHEET 2: Attended Participants
    # -------------------------------------------------------------
    ws_attended = wb.create_sheet(title="Attended Attendees")
    ws_attended.views.sheetView[0].showGridLines = True
    
    headers_attended = ["#", "Ticket Code", "Full Name", "Email", "Phone", "Organization", "Role", "Check-in Timestamp", "Type"]
    for cf in custom_fields:
        headers_attended.append(cf.get("label", cf.get("id")))

    for col_idx, h in enumerate(headers_attended, start=1):
        cell = ws_attended.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    row_num = 2
    for idx, a in enumerate([x for x in attendees if x.get("status") == "Attended"], start=1):
        cdata = {}
        if a.get("custom_data"):
            try:
                cdata = json.loads(a["custom_data"]) if isinstance(a["custom_data"], str) else a["custom_data"]
            except Exception:
                cdata = {}

        row_data = [
            idx,
            a.get("ticket_code", ""),
            a.get("full_name", ""),
            a.get("email", ""),
            a.get("phone", ""),
            a.get("organization", ""),
            a.get("role", ""),
            a.get("checkin_at", ""),
            "Walk-in" if a.get("is_walkin") else "Pre-Registered"
        ]
        for cf in custom_fields:
            row_data.append(str(cdata.get(cf.get("id"), "")))

        for c_idx, val in enumerate(row_data, start=1):
            cell = ws_attended.cell(row=row_num, column=c_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if c_idx in [1, 2, 8, 9]:
                cell.alignment = Alignment(horizontal="center")
        row_num += 1

    for col in ws_attended.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_attended.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # -------------------------------------------------------------
    # SHEET 3: No-Shows & Follow-Up Log
    # -------------------------------------------------------------
    ws_absent = wb.create_sheet(title="Absentee Follow-Ups")
    ws_absent.views.sheetView[0].showGridLines = True
    
    # Map follow-up by attendee_id
    fup_map = {f.get("attendee_id"): f for f in followups}
    
    headers_absent = ["#", "Ticket Code", "Full Name", "Email", "Phone", "Organization", "Follow-up Status", "Channel", "Reason Category", "Organizer Notes"]
    for cf in custom_fields:
        headers_absent.append(cf.get("label", cf.get("id")))

    for col_idx, h in enumerate(headers_absent, start=1):
        cell = ws_absent.cell(row=1, column=col_idx, value=h)
        cell.fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid") # Dark Red
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    row_num = 2
    absent_list = [x for x in attendees if x.get("status") == "Registered"]
    for idx, a in enumerate(absent_list, start=1):
        fup = fup_map.get(a.get("id"), {})
        cdata = {}
        if a.get("custom_data"):
            try:
                cdata = json.loads(a["custom_data"]) if isinstance(a["custom_data"], str) else a["custom_data"]
            except Exception:
                cdata = {}

        row_data = [
            idx,
            a.get("ticket_code", ""),
            a.get("full_name", ""),
            a.get("email", ""),
            a.get("phone", ""),
            a.get("organization", ""),
            fup.get("status", "Not Contacted"),
            fup.get("channel", "WhatsApp"),
            fup.get("reason_category", "Not Specified"),
            fup.get("notes", "")
        ]
        for cf in custom_fields:
            row_data.append(str(cdata.get(cf.get("id"), "")))

        for c_idx, val in enumerate(row_data, start=1):
            cell = ws_absent.cell(row=row_num, column=c_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if c_idx in [1, 2, 7, 8]:
                cell.alignment = Alignment(horizontal="center")
        row_num += 1

    for col in ws_absent.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_absent.column_dimensions[col_letter].width = max(max_len + 3, 14)

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
